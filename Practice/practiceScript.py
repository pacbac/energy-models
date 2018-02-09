from openpyxl import load_workbook

def main():
    wb = load_workbook("data.xlsx")
    sheet = wb.active
    table = {}
    for c in range(4, sheet.max_column):
        name = sheet.cell(row=1, column=c).value
        table[name] = {}
        for r in range(1, sheet.max_row):
            objProp = sheet.cell(row=r, column=c).value
            table[name][objProp] = sheet.cell(row=r, column=c).value
    print(table["Amridge University"]["CITY"])


if __name__ == "__main__":
    main()
