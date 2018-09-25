from openpyxl import load_workbook
from openpyxl import Workbook
from globalFuncs import *

def main():
    wb1 = load_workbook("ProblemCData.xlsx")
    wb3 = load_workbook("CPIAZ.xlsx")
    cpiSheet = wb3.active

    dataSheet = wb1.worksheets[0]
    msnCodes = wb1.worksheets[1]
    table = {}
    msnTable = {}
    GDPRatioTable = [{}, {}, {}, {}]
    inflationTable = [{}, {}, {}, {}]
    AZTable = {}

    for r in range(2, dataSheet.max_row+1):
        msn = dataSheet.cell(row=r, column=1).value #col 1
        state = stateToIndex(dataSheet.cell(row=r, column=2).value) #col 2
        yr = int(dataSheet.cell(row=r, column=3).value) #col 3
        data = float(dataSheet.cell(row=r, column=4).value) #col 4
        if msn not in table:
            table[msn] = [{}, {}, {}, {}] #0=AZ, 1=CA, 2=NM, 3=TX
        table[msn][state][yr] = data

    for r in range(2, msnCodes.max_row+1):
        msn = msnCodes.cell(row=r, column=1).value
        desc = msnCodes.cell(row=r, column=2).value
        unit = msnCodes.cell(row=r, column=3).value
        msnTable[msn] = [desc, unit] #array of [description, unit of msn]

    for r in range(1, cpiSheet.max_row+1):
        AZTable[1996+r] = cpiSheet.cell(row=r, column=3).value

    for state in range(0, 4):
        for yr in table["GDPRX"][state].keys():
            GDPRatioTable[state][yr] = float(table["GDPRV"][state][yr]/table["GDPRX"][state][yr])
        for yr in range(1978, 2010):
            inflationTable[state][yr] = float((GDPRatioTable[state][yr]-GDPRatioTable[state][yr-1])/GDPRatioTable[state][yr-1])
    wb2 = Workbook()
    wb2Sheet = wb2.active

    wb2Sheet.cell(row=1, column=1).value = "Year"
    wb2Sheet.cell(row=1, column=2).value = "CPI"

    yr = 1997
    r = 2
    inflationIndex = float((GDPRatioTable[0][yr] - GDPRatioTable[0][yr-1])/GDPRatioTable[0][yr-1])
    cpiPrev = AZTable[yr]/(1+inflationIndex)
    print(inflationIndex, cpiPrev)
    while yr > 1978:
        wb2Sheet.cell(row=r, column=1).value = yr-1
        wb2Sheet.cell(row=r, column=2).value = cpiPrev
        yr -= 1
        inflationIndex = float((GDPRatioTable[0][yr] - GDPRatioTable[0][yr-1])/GDPRatioTable[0][yr-1])
        cpiPrev = cpiPrev/(1+inflationIndex)
        r += 1

    wb2.save("cpiConsumption"+indexToState(0)+".xlsx")
    pass

if __name__ == "__main__":
    main()
