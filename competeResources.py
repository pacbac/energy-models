import econModel
from econModel import *
from rungeKutta import *
from realGDP import *
from globalFuncs import *
from variableFuncs import *
from openpyxl import Workbook
import math

competeDict = {}
stateNum = 2
#key: type, val: renewable?, THIS LIST DOES NOT INCLUDE OTHER RENEWABLE ENERGY RESOURCES
ourKeys = {"RFTCB": False, "POTCB": False, "NGTCB": False, "MMTCB": False, "LGTCB": False, "JFTCB": False, "HYTCB": True, "CLKCB": False, "CLTXB": False, "BMTCB": True}
otherRenewableKeys = ["GETCB", "SOTCB"] #wind energy excluded b/c it produces energy

r = 2
#Write Model Data to a new worksheet
wb = Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Year"
sheet.cell(row=1, column=3).value = "Total Average Price"
sheet.cell(row=1, column=4).value = "Renewable Average Average"
sheet.cell(row=1, column=5).value = "Description"

def load():
    for key in ourKeys.keys():
        if key not in competeDict:
            competeDict[key] = {yr: table[key][stateNum][yr] for yr in table["GDPRX"][stateNum].keys()}

    competeDict["ORETCB"] = {}
    for yr in table["GDPRX"][stateNum].keys():
        sumOtherRenewables = 0
        for i in otherRenewableKeys:
            sumOtherRenewables += table[i][stateNum][yr]
        competeDict["ORETCB"][yr] = sumOtherRenewables


def calcF(bigA, k, bigQ, sigma): #calc renewable & nonRenewable from NList
    return bigA*k**(sigma)*bigQ**(1-sigma)

def combine(yr, state, NList, currentF, sigma):
    k = 0
    bigQ = 0
    smallq = calcRenewableAvgPrice(yr, state)
    bigA = calcProportion(yr, state)
    for key in NList:
        if key in ourKeys and ourKeys[key]:
            bigQ += NList[key].y
        elif key in ourKeys and not ourKeys[key]:
            k += NList[key].y
        else:
            bigQ += NList[key].y
    p = calcPTotal(yr, state, bigQ, currentF)
    if yr == 2025 or yr == 2050:
        print("yr", yr)
        print("smallq", str(smallq))
        print("p", str(p))
    sigma = calcSigma(bigQ, smallq, p, currentF)
    nextF = calcF(bigA, k, bigQ, sigma)
    coordList = {key: Coord(yr, NList[key].y) for key in NList.keys()}
    return [nextF, coordList]

def splitToNList(yr, endYr, prevF, currentF, coordList, state, prevNList, sigma):
    global r
    NList = {key: approx(0, 1, coordList[key], 1, yr, prevF, key, prevNList, sigma) for key in prevNList.keys()} #set up N(t = start)
    for key in NList:
        if key != "ORETCB":
            sheet.cell(row=r, column=5).value = msnTable[key][0].replace("total consumption", "")
        else:
            sheet.cell(row=r, column=5).value = "Other renewable energy consumption"
        r+=1
    if yr == endYr:
        return NList
    combineArr = combine(yr, state, NList, currentF, sigma)
    yr+=1
    return splitToNList(yr, endYr, currentF, combineArr[0], combineArr[1], state, prevNList, sigma)

def predict(yr, NList, currentTConsump, endYr, state):
    k = 0
    bigQ = 0
    smallq = calcRenewableAvgPrice(yr, state)
    bigA = calcProportion(yr, state)
    for key in NList:
        if key in ourKeys and ourKeys[key]:
            bigQ += NList[key]
        else:
            k += NList[key]
    sumOtherRenewables = 0
    for i in otherRenewableKeys:
        sumOtherRenewables += table[i][stateNum][yr]
    bigQ += sumOtherRenewables
    print(bigQ/currentTConsump)
    p = calcPTotal(yr, state, bigQ, currentTConsump)
    sigma = calcSigma(bigQ, smallq, p, currentTConsump)
    nextF = calcF(bigA, k, bigQ, sigma)
    coordList = {key: Coord(yr, NList[key]) for key in NList.keys()}
    yr+=1
    return splitToNList(yr, endYr, currentTConsump, nextF, coordList, state, NList, sigma)

def main():
    global stateNum
    global table
    global wb
    startYear = 2009 #1978 <= year <= 2009
    endYear = 2051

    for stateNum in range(0, 4):
        print(indexToState(stateNum), "==============================================")
        state = stateNum
        NList = {key: table[key][state][startYear] for key in ourKeys.keys()} #set up N(t = start)
        NList["POTCB"] += table["PATCB"][state][startYear] - table["RFTCB"][state][startYear] - table["MGTCB"][state][startYear] - table["LGTCB"][state][startYear] - table["JFTCB"][state][startYear]
        NList["ORETCB"] = competeDict["ORETCB"][startYear]
        print(NList)
        currentTConsump = table["TETCB"][state][startYear]
        print("currentTConsump", currentTConsump)
        newList = predict(startYear, NList, currentTConsump, endYear, state)
        print("NEWLIST KEYS:")
        for key in newList.keys():
            print(key, newList[key].y)
        stateNum += 1
    pass

load()
if __name__ == "__main__":
    main()
