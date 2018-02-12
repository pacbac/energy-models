from openpyxl import load_workbook
from globalFuncs import *

def main():
    wb = load_workbook("ProblemCData.xlsx")
    sheet = wb.worksheets[0]
    table = {}

    for r in range(2, sheet.max_row+1):
        msn = sheet.cell(row=r, column=1).value #col 1
        state = stateToIndex(sheet.cell(row=r, column=2).value) #col 2
        yr = int(sheet.cell(row=r, column=3).value) #col 3
        data = float(sheet.cell(row=r, column=4).value) #col 4
        if msn not in table:
            table[msn] = [{}, {}, {}, {}] #0=AZ, 1=CA, 2=NM, 3=TX
        table[msn][state][yr] = data

    while(True):
        msn = input("Enter MSN: ")
        #print(type(msn))
        state = stateToIndex(input("Enter state: "))
        #print(type(state))
        yr = int(input("Enter yr: "))
        #print(type(yr))
        try:
            print(table[msn][state][yr])
        except:
            print("Data not found.")
        print()
    pass

if __name__ == "__main__":
    main()
